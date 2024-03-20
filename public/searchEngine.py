import json
import os
import PyPDF2
import logging
from openpyxl import load_workbook
from elasticsearch import Elasticsearch
# logging.basicConfig(level=logging.DEBUG)
# Initialize Elasticsearch connection
es = Elasticsearch([{'host': '127.0.0.1', 'port': 9200, 'scheme': 'http'}])


def search_in_pdf(file_path, keyword):
    with open(file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            index_document(file_path, text)
            if keyword in text:
                return True
    return False


def search_in_excel(file_path, keyword):
    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if keyword in str(cell_value):
                    return True
    return False

def index_document(file_path, content):
    document = {
        'file_path': file_path,
        'content': content,
    }
    es.index(index='your_index', body=document)

def deep_search(directory, keyword):
    matching_files = []
    files = os.listdir(directory)
    
    for file in files:
        file_path = os.path.join(directory, file)
        if file_path.endswith('.pdf') and search_in_pdf(file_path, keyword):
            matching_files.append(file)
            
        elif file_path.endswith('.xlsx') and search_in_excel(file_path, keyword):
            matching_files.append(file)     
    return matching_files

if __name__ == "__main__":
    import sys 
    keyword = sys.argv[1]
    result = deep_search("\\storage\\app\\public\\", keyword)
    result_json = json.dumps(result)
    print(result_json)
